"""XLSX text extraction."""

from __future__ import annotations

from openpyxl import load_workbook


def parse_xlsx(file_path: str) -> str:
    """Return pipe-separated text of all sheets in a .xlsx file."""
    workbook = load_workbook(filename=file_path, data_only=True)
    sheets: list[str] = []
    for sheet in workbook.worksheets:
        lines = [f"Sheet: {sheet.title}"]
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                lines.append(" | ".join(cells))
        sheets.append("\n".join(lines))
    return "\n\n".join(sheets)
